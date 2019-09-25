import numpy as np
import openpyxl as xl
from matplotlib import pyplot as plt

def get_size():
    print("Введите размеры матрицы  : ")
    try:
        m, n = map(int, input().split())
    except:
        m = 0
        n = 0
    return m, n

def create_mtr(m, n):
    mtr = np.random.rand(np.abs(m), np.abs(n))
    mtr *= 20
    mtr -= 10
    return mtr

def save_mtr(mtr, num, b):
    if mtr.shape != 0:
        spam = b.sheetnames
        if spam[0] == "Sheet": b.remove(b[spam[0]])
        sh = b.create_sheet("Матрица " +str(num))
        for i in list(mtr):
            sh.append(list(i))

def sum_mtr(mtr1, mtr2):
    if mtr1.shape == mtr2.shape:
        return mtr1 + mtr2, 1
    else:
        return 0, 0

def multi_mtr(mtr1, mtr2):
    if mtr1.shape[1] == mtr2.shape[0]:
        return mtr1 @ mtr2, 1
    else:
        return 0, 0

m1, n1 = get_size()
m2, n2 = get_size()
if (m1 * n1 * m2 * n2 != 0):
    mtr1 = create_mtr(m1, n1)
    mtr2 = create_mtr(m2, n2)
    b = xl.Workbook()
    filename = "results.xlsx"
    save_mtr(mtr1, 1, b)
    save_mtr(mtr2, 2, b)
    #Сумма и умножение
    mtr3, k = sum_mtr(mtr1, mtr2)
    if k == 1: save_mtr(mtr3, 3, b)
    mtr4, r = multi_mtr(mtr1, mtr2)
    if r == 1: save_mtr(mtr4, 4, b)
    b.save(filename)
    #Создание графика
    f = plt.figure()
    f1 = f.add_subplot(4,1,1)
    f1.plot(mtr1.sum(axis=0),"-b")
    f1.plot(mtr1.sum(axis=0),"ro")

    f2 = f.add_subplot(4,1,2)
    f2.plot(mtr2.sum(axis=0), "-g")
    f2.plot(mtr2.sum(axis=0), "ro")
    if k==1:
        f3 = f.add_subplot(4, 1, 3)
        f3.plot(mtr3.sum(axis=0), "-y")
        f3.plot(mtr3.sum(axis=0), "ro")
    if r==1:
        f4 = f.add_subplot(4, 1, 4)
        f4.plot(mtr4.sum(axis=0), "-o")
        f4.plot(mtr4.sum(axis=0), "ro")
        #plt.bar(np.arange())
    plt.show()

else:
    print("Невозможно создать одну из матриц")
